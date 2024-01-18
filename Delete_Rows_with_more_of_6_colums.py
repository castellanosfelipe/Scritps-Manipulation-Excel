# -*- coding: utf-8 -*-
"""
Created on Thu Jan 18 08:52:56 2024

@author: bpena
"""

import openpyxl


archivo_excel = "tu_archivo.xlsx"  # Reemplaza con el nombre de tu archivo
workbook = openpyxl.load_workbook(archivo_excel)
sheet = workbook.active  

filas_a_copiar = []

# Recorre las celdas de la columna F y registra las filas a copiar
for idx, row in enumerate(sheet.iter_rows(min_col=6, max_col=6, min_row=2, values_only=True), start=2):
    valor_celda = row[0]
    if valor_celda is not None and valor_celda.startswith("Anexo"):
        filas_a_copiar.append(idx)

# Copia los valores de las filas a la columna H y luego elimina las celdas copiadas de la columna F
for fila_a_copiar in reversed(filas_a_copiar):
    valor_celda = sheet.cell(row=fila_a_copiar, column=6).value
    sheet.cell(row=fila_a_copiar, column=8).value = valor_celda
    sheet.cell(row=fila_a_copiar, column=6).value = None

# Elimina las filas que tienen menos de 6 columnas con información
filas_a_eliminar = [idx for idx, row in enumerate(sheet.iter_rows(min_row=2), start=2) if len([cell for cell in row if cell.value is not None]) < 6]
for fila_a_eliminar in reversed(filas_a_eliminar):
    sheet.delete_rows(fila_a_eliminar)
    
# Guarda los cambios en el archivo Excel
workbook.save(archivo_excel)
