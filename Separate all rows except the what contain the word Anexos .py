# -*- coding: utf-8 -*-
"""
Created on Thu Jan 18 09:14:32 2024

@author: bpena
"""

import openpyxl
import os

# Cargar el archivo Excel de entrada
archivo_entrada = "archivo_entrada.xlsx"
workbook = openpyxl.load_workbook(archivo_entrada)
sheet = workbook.active

for row in sheet.iter_rows():
    for cell in row:
        if "\\" in str(cell.value):
            directorios = str(cell.value).split("\\")
            
            # Buscar la posición de la palabra "Anexo"
            if "Anexo" in directorios:
                index_anexo = directorios.index("Anexo")
            else:
                index_anexo = len(directorios)  # Si no se encuentra "Anexo", tomar toda la cadena
            
            cell.value = None
            
            for i in range(index_anexo):
                sheet.cell(row=cell.row, column=cell.column + i).value = directorios[i]
            
            if index_anexo < len(directorios):
                sheet.cell(row=cell.row, column=cell.column + index_anexo).value = "\\".join(directorios[index_anexo:])

filas_a_copiar = []

for idx, row in enumerate(sheet.iter_rows(min_col=6, max_col=6, min_row=2, values_only=True), start=2):
    valor_celda = row[0]
    if valor_celda is not None and valor_celda.startswith("Anexo"):
        filas_a_copiar.append(idx)

# Copia los valores de las filas a la columna H y luego elimina las celdas copiadas de la columna F
for fila_a_copiar in reversed(filas_a_copiar):
    valor_celda = sheet.cell(row=fila_a_copiar, column=6).value
    sheet.cell(row=fila_a_copiar, column=8).value = valor_celda
    sheet.cell(row=fila_a_copiar, column=6).value = None

# Elimina las filas que tienen menos de 6 columnas con información
filas_a_eliminar = [idx for idx, row in enumerate(sheet.iter_rows(min_row=2), start=2) if len([cell for cell in row if cell.value is not None]) < 6]
for fila_a_eliminar in reversed(filas_a_eliminar):
    sheet.delete_rows(fila_a_eliminar)

# Recorrer las filas de la columna H y copiar el último segmento del directorio a la columna I
for fila in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=8, max_col=8):  # Columna H
    directorio = fila[0].value
    if directorio:
        segmentos = directorio.split(os.path.sep)  
        ultimo_segmento = segmentos[-1] if segmentos else "" 
        fila[0].offset(column=1).value = ultimo_segmento  # Copiarlo segmento a la columna I

# Guardar el archivo Excel modificado
archivo_salida = "archivo_salida.xlsx"
workbook.save(archivo_salida)

# Cerrar el archivo Excel
workbook.close()
