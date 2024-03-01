# -*- coding: utf-8 -*-
"""
Created on Thu Feb 29 16:07:57 2024

@author: bpena
"""

import openpyxl

# Abre el archivo Excel
archivo_excel = "Li.xlsx"
libro_trabajo = openpyxl.load_workbook(archivo_excel)
hoja = libro_trabajo.active

# Recorre cada celda en la columna A
for fila in hoja.iter_rows(min_row=1, max_col=1, max_row=hoja.max_row):
    for celda in fila:
        # Divide la ruta en segmentos usando '\' como separador y toma el último segmento
        ultimo_segmento = celda.value.split("\\")[-1]
        
        # Actualiza el valor de la celda con el último segmento de la ruta
        celda.value = ultimo_segmento

# Guarda los cambios en un nuevo archivo Excel
libro_trabajo.save("nuevo_archivo.xlsx")

