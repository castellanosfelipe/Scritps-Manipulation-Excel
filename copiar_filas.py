# -*- coding: utf-8 -*-
"""
Created on Fri Jan 12 16:05:34 2024

@author: bpena
"""
import openpyxl

def comparar_y_copiar_excel(ruta_origen, ruta_destino):
    # Cargar los archivos Excel
    libro_origen = openpyxl.load_workbook(ruta_origen)
    libro_destino = openpyxl.load_workbook(ruta_destino)

    # Obtener las hojas de los libros
    hoja_origen = libro_origen.active
    hoja_destino = libro_destino.active

    # Identificar las filas que están en el archivo de destino pero no en el de origen
    filas_nuevas = []
    for fila_destino in hoja_destino.iter_rows(min_row=2, values_only=True):
        existe_en_origen = False
        for fila_origen in hoja_origen.iter_rows(min_row=2, values_only=True):
            if fila_destino[:-1] == fila_origen[:-1]:  # Compara todas las columnas excepto la última
                existe_en_origen = True
                break

        if not existe_en_origen:
            filas_nuevas.append(fila_destino)

    # Copiar las filas nuevas al archivo de origen conservando el formato
    for fila_nueva in filas_nuevas:
        hoja_origen.append(fila_nueva)
        # Obtener la celda recién añadida y la celda correspondiente en la hoja de destino
        celda_origen = hoja_origen.cell(row=hoja_origen.max_row, column=hoja_destino.max_column)
        celda_destino = hoja_destino.cell(row=hoja_destino.max_row, column=hoja_destino.max_column)
        # Copiar el estilo de la celda de destino a la celda de origen
        celda_origen.font = celda_destino.font.copy()
        celda_origen.border = celda_destino.border.copy()
        celda_origen.fill = celda_destino.fill.copy()
        celda_origen.number_format = celda_destino.number_format  # No se utiliza copy() en number_format
        celda_origen.protection = celda_destino.protection.copy()
        celda_origen.alignment = celda_destino.alignment.copy()

    # Guardar el resultado de vuelta al archivo de origen
    libro_origen.save(ruta_origen)
    print("Se han copiado las filas nuevas al archivo de origen conservando el formato.")

# Rutas de los archivos Excel
ruta_destino = "C:/c/Horas_soporte.xlsx"
ruta_origen = "C:/c/datos.xlsx"

# Llamar a la función
comparar_y_copiar_excel(ruta_origen, ruta_destino)


